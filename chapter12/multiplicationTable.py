#!/usr/bin/env python
import openpyxl, sys

def main():
    wb = openpyxl.Workbook()
    sheet = wb.active

    for index in range(1, int(sys.argv[1]) + 1):
        sheet.cell(row=index + 1, column=1).value = index
        sheet.cell(row=1, column=index + 1).value = index

    for row in range(2, sheet.max_row + 1):
        for col in range(2, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = (row - 1) * (col - 1)

    wb.save('multiplication.xlsx')

if __name__ == '__main__':
    main()
